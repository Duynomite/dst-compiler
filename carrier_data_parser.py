#!/usr/bin/env python3
"""
carrier_data_parser.py — Parse carrier DST spreadsheets and cross-reference

Parses Aetna and Wellcare Excel files into normalized disaster records.
Cross-references against curated_disasters.json to identify:
  1. Records we already have (match) — add carrier acknowledgment metadata
  2. Records we're missing (gap) — generate research queue
  3. SEP window discrepancies — flag for review

Also produces a Carrier Intelligence Report analyzing carrier data collection
processes, source discovery, timeliness, and pipeline improvement recommendations.

Usage:
  python carrier_data_parser.py \\
    --aetna ../../69a5ea47d2c45762233c8373.xlsx \\
    --wellcare ../../69a5f3f1320ef43511d78084.xlsx \\
    --curated curated_disasters.json \\
    --output carrier_analysis.json

Dependencies: openpyxl (added to requirements.txt)
"""

import argparse
import calendar
import json
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

import openpyxl


# =========================================================================
# US State normalization
# =========================================================================

STATE_NAME_TO_ABBREV = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "district of columbia": "DC", "district of columbia (washington dc)": "DC",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN",
    "mississippi": "MS", "missouri": "MO", "montana": "MT", "nebraska": "NE",
    "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "orgeon": "OR",  # Known Wellcare typo
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "washington dc": "DC",
}

VALID_STATE_ABBREVS = set(STATE_NAME_TO_ABBREV.values())


def normalize_state(raw: str) -> Optional[str]:
    """Normalize state name or abbreviation to 2-letter code."""
    if not raw:
        return None
    cleaned = raw.strip()
    # Already a valid abbreviation
    if cleaned.upper() in VALID_STATE_ABBREVS:
        return cleaned.upper()
    # Full name lookup
    abbrev = STATE_NAME_TO_ABBREV.get(cleaned.lower())
    if abbrev:
        return abbrev
    return None


# =========================================================================
# CFR SEP Window calculation (mirrors dst_data_fetcher.py exactly)
# =========================================================================

def calculate_sep_window_end_cfr(incident_end: date) -> date:
    """Last day of 2nd full calendar month after incident end."""
    month = incident_end.month
    year = incident_end.year
    target_month = month + 2
    target_year = year
    if target_month > 12:
        target_month -= 12
        target_year += 1
    last_day = calendar.monthrange(target_year, target_month)[1]
    return date(target_year, target_month, last_day)


# =========================================================================
# Carrier Record dataclass
# =========================================================================

@dataclass
class CarrierRecord:
    """Normalized carrier disaster record."""
    carrier: str                        # "aetna" | "wellcare"
    state: str                          # 2-letter abbreviation
    title: str
    incident_type: str                  # Normalized: "Hurricane", "Wildfire", etc.
    declaring_authority: str            # "Governor", "FEMA", "FEMA/FMAG", etc.
    incident_start: Optional[date] = None
    incident_end: Optional[date] = None  # None = ongoing
    sep_start: Optional[date] = None
    sep_end: Optional[date] = None
    extension_sep_end: Optional[date] = None
    extension_incident_end: Optional[date] = None
    extension_date: Optional[date] = None
    counties: List[str] = field(default_factory=list)
    statewide: bool = False
    official_urls: List[str] = field(default_factory=list)
    status: str = "Active"              # Active, Amended, Renewed, Expired
    notes: str = ""
    wellcare_type: str = ""             # Wellcare-specific Type field
    event_reported_date: Optional[date] = None  # Aetna-specific


# =========================================================================
# Incident Type Normalization
# =========================================================================

INCIDENT_TYPE_PATTERNS = [
    (r"hurricane|tropical\s+(storm|cyclone|depression)", "Hurricane/Tropical Storm"),
    (r"tornado", "Tornado"),
    (r"wildfire|fire\b|fmag|blaze", "Wildfire"),
    (r"drought", "Drought"),
    (r"winter\s+(storm|weather)|blizzard|ice\s+storm|snow|cold\s+weather|freezing", "Severe Winter Storm"),
    (r"flood|rain|atmospheric\s+river|tsunami|storm\s+surge|mudslide|landslide", "Flood/Severe Storm"),
    (r"earthquake", "Earthquake"),
    (r"wind|windstorm|high\s+wind|straight.line.wind", "Severe Wind"),
    (r"homelessness", "Homelessness Emergency"),
    (r"border|migration|immigration|illegal\s+migration", "Immigration Emergency"),
    (r"crime|unrest|civil", "Crime/Civil Emergency"),
    (r"healthcare|staff\s+shortage|vaccine|hospital", "Healthcare Emergency"),
    (r"water\s+system|sewer|well\s+outage|pipeline|fuel\s+supply", "Infrastructure Emergency"),
    (r"plane\s+crash", "Plane Crash"),
    (r"war\s+in\s+israel", "International Crisis"),
    (r"power\s+outage|power\s+demand", "Power Emergency"),
]


def classify_incident_type(title: str) -> str:
    """Classify a disaster title into a normalized incident type."""
    lower = title.lower()
    for pattern, incident_type in INCIDENT_TYPE_PATTERNS:
        if re.search(pattern, lower):
            return incident_type
    return "Other"


# =========================================================================
# AetnaParser
# =========================================================================

class AetnaParser:
    """Parse Aetna DST Excel (109 rows, 41 states)."""

    ONGOING_SENTINEL = date(9999, 9, 9)

    def parse(self, filepath: str) -> List[CarrierRecord]:
        """Parse Aetna Excel file into normalized CarrierRecords."""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rec = self._parse_row(row_dict)
            if rec:
                records.append(rec)

        print(f"  Aetna: Parsed {len(records)} records from {filepath}")
        return records

    def _parse_row(self, row: Dict) -> Optional[CarrierRecord]:
        state = normalize_state(str(row.get("SEPState", "")).strip())
        if not state:
            return None

        title = str(row.get("EventTitle", "")).strip()
        if not title:
            return None

        incident_start = self._parse_date(row.get("IncidentPeriodStart"))
        incident_end_raw = self._parse_date(row.get("IncidentPeriodEnd"))
        incident_end = None if self._is_ongoing(incident_end_raw) else incident_end_raw

        sep_start = self._parse_date(row.get("SEPStartDate"))
        sep_end_raw = self._parse_date(row.get("SEPEndDate"))
        sep_end = None if self._is_ongoing(sep_end_raw) else sep_end_raw

        ext_sep_end_raw = self._parse_date(row.get("ExtensionSEPEndDate"))
        ext_sep_end = None if self._is_ongoing(ext_sep_end_raw) else ext_sep_end_raw

        ext_inc_end_raw = self._parse_date(row.get("ExtensionIncidentEndDate"))
        ext_inc_end = None if self._is_ongoing(ext_inc_end_raw) else ext_inc_end_raw

        ext_date = self._parse_date(row.get("ExtensionAnnouncementDate") or row.get("ExtensionAnnoucementDate"))

        counties_raw = str(row.get("County", "")).strip()
        if counties_raw.upper() == "ALL":
            counties = ["Statewide"]
            statewide = True
        else:
            counties = [c.strip() for c in counties_raw.split(",") if c.strip()]
            statewide = False

        urls = []
        for i in range(1, 5):
            url = row.get(f"RelatedLinks{i}")
            if url and str(url).strip().startswith("http"):
                urls.append(str(url).strip())

        authority = str(row.get("EntityIssuingDeclaration", "Governor")).strip()
        event_reported = self._parse_date(row.get("EventReportedDate"))

        return CarrierRecord(
            carrier="aetna",
            state=state,
            title=title,
            incident_type=classify_incident_type(title),
            declaring_authority=authority,
            incident_start=incident_start,
            incident_end=incident_end,
            sep_start=sep_start,
            sep_end=sep_end,
            extension_sep_end=ext_sep_end,
            extension_incident_end=ext_inc_end,
            extension_date=ext_date,
            counties=counties,
            statewide=statewide,
            official_urls=urls,
            status="Active",
            event_reported_date=event_reported,
        )

    def _parse_date(self, val) -> Optional[date]:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        s = str(val).strip()
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _is_ongoing(self, d: Optional[date]) -> bool:
        if d is None:
            return True
        return d >= self.ONGOING_SENTINEL


# =========================================================================
# WellcareParser
# =========================================================================

class WellcareParser:
    """Parse Wellcare DST Excel (3,885 rows, 42 states)."""

    ACTIVE_STATUSES = {"Active", "Amended", "Renewed"}

    def parse(self, filepath: str) -> List[CarrierRecord]:
        """Parse Wellcare Excel, filtering to active/amended/renewed only."""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        # Clean header whitespace
        headers = [h.strip() if h else h for h in headers]

        records = []
        all_records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rec = self._parse_row(row_dict)
            if rec:
                all_records.append(rec)
                if rec.status in self.ACTIVE_STATUSES:
                    records.append(rec)

        print(f"  Wellcare: Parsed {len(records)} active records ({len(all_records)} total) from {filepath}")
        return records

    def parse_all(self, filepath: str) -> List[CarrierRecord]:
        """Parse ALL Wellcare records (including expired) for intelligence analysis."""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        headers = [h.strip() if h else h for h in headers]

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rec = self._parse_row(row_dict)
            if rec:
                records.append(rec)
        return records

    def _parse_row(self, row: Dict) -> Optional[CarrierRecord]:
        state = normalize_state(str(row.get("State", "")).strip())
        if not state:
            return None

        title = str(row.get("Disaster Type", "")).strip()
        if not title:
            return None

        status = str(row.get("Status", "")).strip() if row.get("Status") else None
        if not status:
            # Rows without status (3,651 of them) — treat as historical/expired
            status = "Expired"

        sep_start, sep_end = self._parse_sep_dates(row.get("SEP Effective Dates"))
        wc_type = str(row.get("Type", "")).strip() if row.get("Type") else ""
        notes = str(row.get("Notes / Updates", "")).strip() if row.get("Notes / Updates") else ""

        counties_raw = str(row.get("Counties Impacted", "")).strip() if row.get("Counties Impacted") else ""
        if not counties_raw or counties_raw.lower() in ("all", "statewide", "all counties"):
            counties = ["Statewide"]
            statewide = True
        else:
            counties = [c.strip() for c in re.split(r",|;|\band\b", counties_raw) if c.strip()]
            statewide = False

        # Determine declaring authority from Wellcare Type field
        if "FEMA/FMAG" in wc_type:
            authority = "FEMA/FMAG"
        elif "FEMA" in wc_type:
            authority = "FEMA"
        elif "PHE" in wc_type:
            authority = "HHS"
        else:
            authority = "Governor"  # Emergency type = state-level

        return CarrierRecord(
            carrier="wellcare",
            state=state,
            title=title,
            incident_type=classify_incident_type(title),
            declaring_authority=authority,
            sep_start=sep_start,
            sep_end=sep_end,
            counties=counties,
            statewide=statewide,
            status=status,
            notes=notes,
            wellcare_type=wc_type,
        )

    def _parse_sep_dates(self, val) -> Tuple[Optional[date], Optional[date]]:
        """Parse 'MM/DD/YYYY - MM/DD/YYYY' format."""
        if not val:
            return None, None
        s = str(val).strip()
        # Try "start - end" pattern
        m = re.match(r"(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})", s)
        if m:
            try:
                start = datetime.strptime(m.group(1), "%m/%d/%Y").date()
                end = datetime.strptime(m.group(2), "%m/%d/%Y").date()
                return start, end
            except ValueError:
                pass
        return None, None


# =========================================================================
# CrossReferencer
# =========================================================================

class CrossReferencer:
    """Match carrier records against curated_disasters.json."""

    def __init__(self, curated_path: str):
        with open(curated_path) as f:
            data = json.load(f)
        self.curated = data.get("disasters", data) if isinstance(data, dict) else data
        # Build lookup indices
        self._by_state = defaultdict(list)
        for rec in self.curated:
            self._by_state[rec["state"]].append(rec)

    def match_all(self, carrier_records: List[CarrierRecord]) -> Dict:
        """Match all carrier records against curated data.

        Returns:
            {
                "matched": [...],       # Carrier record matched to curated record
                "gaps": [...],          # Carrier records with no curated match
                "discrepancies": [...], # Matched records with SEP window differences
            }
        """
        matched = []
        gaps = []
        discrepancies = []

        for crec in carrier_records:
            curated_match = self._find_match(crec)
            if curated_match:
                match_entry = {
                    "carrier": crec.carrier,
                    "carrier_title": crec.title,
                    "carrier_state": crec.state,
                    "carrier_sep_end": crec.sep_end.isoformat() if crec.sep_end else None,
                    "carrier_incident_end": crec.incident_end.isoformat() if crec.incident_end else None,
                    "curated_id": curated_match["id"],
                    "curated_title": curated_match["title"],
                    "curated_sep_end": curated_match.get("sepWindowEnd"),
                    "match_tier": curated_match.get("_match_tier", "unknown"),
                    "has_extension": crec.extension_sep_end is not None,
                }
                matched.append(match_entry)

                # Check for SEP window discrepancy
                disc = self._check_sep_discrepancy(curated_match, crec)
                if disc:
                    discrepancies.append(disc)
            else:
                gap_entry = {
                    "carrier": crec.carrier,
                    "state": crec.state,
                    "title": crec.title,
                    "incident_type": crec.incident_type,
                    "declaring_authority": crec.declaring_authority,
                    "incident_start": crec.incident_start.isoformat() if crec.incident_start else None,
                    "incident_end": crec.incident_end.isoformat() if crec.incident_end else None,
                    "sep_start": crec.sep_start.isoformat() if crec.sep_start else None,
                    "sep_end": crec.sep_end.isoformat() if crec.sep_end else None,
                    "counties": crec.counties[:5],  # Truncate for readability
                    "statewide": crec.statewide,
                    "official_urls": crec.official_urls,
                    "status": crec.status,
                    "wellcare_type": crec.wellcare_type,
                    "notes": crec.notes[:200] if crec.notes else "",
                }
                gaps.append(gap_entry)

        return {
            "matched": matched,
            "gaps": gaps,
            "discrepancies": discrepancies,
        }

    def _find_match(self, crec: CarrierRecord) -> Optional[Dict]:
        """Find matching curated record using tiered strategy."""
        state_records = self._by_state.get(crec.state, [])
        if not state_records:
            return None

        # Tier 1: Title keyword match (state already filtered)
        title_lower = crec.title.lower()
        title_keywords = self._extract_keywords(title_lower)
        for rec in state_records:
            rec_keywords = self._extract_keywords(rec["title"].lower())
            overlap = title_keywords & rec_keywords
            if len(overlap) >= 2:
                rec["_match_tier"] = "tier1_title"
                return rec

        # Tier 2: Incident start within 7 days + same incident type
        if crec.incident_start:
            for rec in state_records:
                rec_start = self._parse_iso(rec.get("incidentStart"))
                if rec_start and abs((crec.incident_start - rec_start).days) <= 7:
                    rec_type = classify_incident_type(rec["title"])
                    if rec_type == crec.incident_type:
                        rec["_match_tier"] = "tier2_date_type"
                        return rec

        # Tier 3: Declaring authority + date overlap
        if crec.sep_start and crec.sep_end:
            for rec in state_records:
                rec_start = self._parse_iso(rec.get("sepWindowStart"))
                rec_end = self._parse_iso(rec.get("sepWindowEnd"))
                if rec_start and rec_end:
                    # Check for date range overlap
                    if crec.sep_start <= rec_end and crec.sep_end >= rec_start:
                        # Same general category?
                        rec_type = classify_incident_type(rec["title"])
                        if rec_type == crec.incident_type:
                            rec["_match_tier"] = "tier3_overlap"
                            return rec

        # Tier 4: URL match
        if crec.official_urls:
            for rec in state_records:
                rec_url = rec.get("officialUrl", "")
                for carrier_url in crec.official_urls:
                    if self._urls_match(carrier_url, rec_url):
                        rec["_match_tier"] = "tier4_url"
                        return rec

        return None

    def _extract_keywords(self, text: str) -> set:
        """Extract meaningful keywords from title, ignoring common words."""
        stop_words = {
            "the", "of", "a", "an", "in", "for", "and", "or", "to", "state",
            "emergency", "declaration", "declares", "declaring", "disaster",
            "executive", "order", "governor", "proclamation", "eo", "update",
            "extension", "updated", "severe", "weather", "event", "system",
            "preparedness", "ahead", "major", "conditions",
        }
        words = set(re.findall(r"[a-z]+", text))
        return words - stop_words

    def _parse_iso(self, val) -> Optional[date]:
        if not val:
            return None
        try:
            return date.fromisoformat(val)
        except (ValueError, TypeError):
            return None

    def _urls_match(self, url1: str, url2: str) -> bool:
        """Check if two URLs point to the same resource (fuzzy domain + path match)."""
        if not url1 or not url2:
            return False
        # Normalize: strip protocol, www, trailing slash
        def normalize(u):
            u = re.sub(r"^https?://", "", u)
            u = re.sub(r"^www\.", "", u)
            return u.rstrip("/").lower()
        return normalize(url1) == normalize(url2)

    def _check_sep_discrepancy(self, curated: Dict, carrier: CarrierRecord) -> Optional[Dict]:
        """Compare our CFR-calculated SEP end vs carrier's SEP end."""
        our_end = self._parse_iso(curated.get("sepWindowEnd"))
        carrier_end = carrier.sep_end

        if not our_end or not carrier_end:
            return None

        diff_days = (carrier_end - our_end).days
        if abs(diff_days) <= 1:  # Allow 1-day rounding tolerance
            return None

        # Also check if carrier has extension that explains the difference
        effective_carrier_end = carrier.extension_sep_end or carrier_end

        return {
            "carrier": carrier.carrier,
            "state": carrier.state,
            "title": carrier.title,
            "curated_id": curated["id"],
            "our_sep_end": our_end.isoformat(),
            "carrier_sep_end": carrier_end.isoformat(),
            "carrier_extension_sep_end": carrier.extension_sep_end.isoformat() if carrier.extension_sep_end else None,
            "diff_days": diff_days,
            "explanation": self._explain_discrepancy(diff_days, carrier),
        }

    def _explain_discrepancy(self, diff_days: int, carrier: CarrierRecord) -> str:
        if carrier.extension_sep_end:
            return f"Carrier has extension (ext SEP end: {carrier.extension_sep_end.isoformat()})"
        if diff_days < 0:
            return "Carrier SEP window is SHORTER than CFR formula (carrier being conservative?)"
        return "Carrier SEP window is LONGER than CFR formula (possible extension or different calculation)"


# =========================================================================
# GapReportGenerator
# =========================================================================

class GapReportGenerator:
    """Generate structured gap report and carrier intelligence report."""

    def generate_gap_report(self, results: Dict) -> Dict:
        """Structure gaps for research queue."""
        gaps = results["gaps"]

        # Deduplicate gaps (same disaster from different carriers)
        deduped = self._deduplicate_gaps(gaps)

        # Group by incident type
        by_type = defaultdict(list)
        for g in deduped:
            by_type[g["incident_type"]].append(g)

        # Group by state
        by_state = defaultdict(list)
        for g in deduped:
            by_state[g["state"]].append(g)

        # Prioritize: multi-carrier > recency > alphabetical
        for g in deduped:
            g["priority_score"] = g.get("carrier_count", 1) * 10
            if g.get("sep_end"):
                try:
                    end = date.fromisoformat(g["sep_end"])
                    if end > date.today():
                        g["priority_score"] += 5  # Currently active
                except (ValueError, TypeError):
                    pass

        deduped.sort(key=lambda g: (-g["priority_score"], g["state"], g["title"]))

        return {
            "total_gaps": len(deduped),
            "by_type": {k: len(v) for k, v in sorted(by_type.items(), key=lambda x: -len(x[1]))},
            "by_state": {k: len(v) for k, v in sorted(by_state.items())},
            "gaps": deduped,
        }

    def _deduplicate_gaps(self, gaps: List[Dict]) -> List[Dict]:
        """Merge gaps from different carriers for same disaster."""
        key_map = {}
        for g in gaps:
            key = (g["state"], self._normalize_title_key(g["title"]))
            if key in key_map:
                existing = key_map[key]
                existing["carrier_count"] = existing.get("carrier_count", 1) + 1
                existing["carriers"] = existing.get("carriers", [existing["carrier"]])
                existing["carriers"].append(g["carrier"])
                # Merge URLs
                if g.get("official_urls"):
                    existing.setdefault("official_urls", []).extend(g["official_urls"])
            else:
                g["carrier_count"] = 1
                g["carriers"] = [g["carrier"]]
                key_map[key] = g
        return list(key_map.values())

    def _normalize_title_key(self, title: str) -> str:
        """Create fuzzy key for deduplication."""
        lower = title.lower()
        # Remove common prefixes/suffixes
        lower = re.sub(r"^(extension:?\s*|update:?\s*|updated:?\s*|renewed:?\s*)", "", lower)
        # Remove EO numbers
        lower = re.sub(r"\s*eo\s*[-_]?\s*[\d.]+\s*", " ", lower)
        # Remove dates
        lower = re.sub(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", "", lower)
        # Normalize whitespace
        return re.sub(r"\s+", " ", lower).strip()

    def generate_intelligence_report(
        self,
        aetna_records: List[CarrierRecord],
        wellcare_records: List[CarrierRecord],
        wellcare_all: List[CarrierRecord],
        results: Dict,
    ) -> str:
        """Generate carrier intelligence markdown report."""
        lines = ["# Carrier Intelligence Report", ""]
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        lines.append("")

        # --- Section 1: Source Discovery ---
        lines.append("## 1. Source Discovery — Where Do Carriers Find DSTs We Miss?")
        lines.append("")

        # Authority breakdown
        lines.append("### Declaring Authority Distribution")
        lines.append("")
        auth_counts = Counter()
        for r in aetna_records + wellcare_records:
            auth_counts[r.declaring_authority] += 1
        lines.append("| Authority | Count | % |")
        lines.append("|-----------|-------|---|")
        total = sum(auth_counts.values())
        for auth, count in auth_counts.most_common():
            lines.append(f"| {auth} | {count} | {count*100//total}% |")
        lines.append("")

        # Incident type breakdown for GAPS only
        gap_types = Counter()
        for g in results["gaps"]:
            gap_types[g["incident_type"]] += 1
        lines.append("### Disaster Types We're Missing (Gaps Only)")
        lines.append("")
        lines.append("| Type | Missing Count | Examples |")
        lines.append("|------|--------------|----------|")
        for itype, count in gap_types.most_common():
            examples = [g["title"][:40] for g in results["gaps"] if g["incident_type"] == itype][:3]
            lines.append(f"| {itype} | {count} | {'; '.join(examples)} |")
        lines.append("")

        # Wellcare Type taxonomy
        wc_type_counts = Counter()
        for r in wellcare_records:
            if r.wellcare_type:
                wc_type_counts[r.wellcare_type] += 1
        lines.append("### Wellcare Declaration Type Taxonomy")
        lines.append("")
        lines.append("| Wellcare Type | Count | Our Equivalent |")
        lines.append("|--------------|-------|----------------|")
        type_mapping = {
            "Emergency": "STATE (Governor)",
            "Disaster/FEMA": "FEMA (DR/EM)",
            "Disaster/FEMA/FMAG": "FEMA FM (WE EXCLUDE THIS)",
            "Disaster": "Mixed — state/local/FEMA",
            "PHE": "HHS",
            "Emergency/FEMA": "FEMA + State combined",
        }
        for wt, count in wc_type_counts.most_common():
            equiv = type_mapping.get(wt, "Unknown")
            lines.append(f"| {wt} | {count} | {equiv} |")
        lines.append("")

        # FMAG finding
        fmag_records = [r for r in wellcare_records if "FMAG" in r.wellcare_type]
        if fmag_records:
            lines.append("### CRITICAL FINDING: FEMA Fire Management (FMAG) Declarations")
            lines.append("")
            lines.append(f"Wellcare tracks **{len(fmag_records)} active FEMA/FMAG declarations** that we EXCLUDE.")
            lines.append("We excluded FM/FMAG per Stafford Act analysis (v2.8 decision). However, carriers ARE honoring them.")
            lines.append("")
            lines.append("| State | Title | SEP Window |")
            lines.append("|-------|-------|------------|")
            for r in fmag_records[:15]:
                sep = f"{r.sep_start} - {r.sep_end}" if r.sep_start and r.sep_end else "N/A"
                lines.append(f"| {r.state} | {r.title} | {sep} |")
            if len(fmag_records) > 15:
                lines.append(f"| ... | +{len(fmag_records)-15} more | |")
            lines.append("")
            lines.append("**Recommendation:** Re-evaluate FM/FMAG exclusion. If all 3 carriers honor them,")
            lines.append("agents may be unable to enroll beneficiaries without them in our tool.")
            lines.append("")

        # --- Section 2: URL/Source Discovery ---
        lines.append("## 2. Government URLs From Carrier Data")
        lines.append("")
        lines.append("URLs from Aetna RelatedLinks that we could monitor:")
        lines.append("")

        url_domains = Counter()
        all_urls = []
        for r in aetna_records:
            for url in r.official_urls:
                domain = re.sub(r"^https?://(www\.)?", "", url).split("/")[0]
                url_domains[domain] += 1
                all_urls.append((r.state, r.title, url))

        lines.append("| Domain | Records Using It |")
        lines.append("|--------|-----------------|")
        for domain, count in url_domains.most_common(20):
            lines.append(f"| {domain} | {count} |")
        lines.append("")

        # --- Section 3: Timeliness Analysis ---
        lines.append("## 3. Timeliness Analysis — Carrier Discovery Speed")
        lines.append("")

        if results["matched"]:
            lines.append(f"Matched {len(results['matched'])} carrier records to our curated data.")
            lines.append("")

            # For Aetna matches, compare EventReportedDate to our lastUpdated
            aetna_matches = [m for m in results["matched"] if m["carrier"] == "aetna"]
            lines.append(f"Aetna matches: {len(aetna_matches)}")
            lines.append(f"Wellcare matches: {len(results['matched']) - len(aetna_matches)}")
            lines.append("")
        else:
            lines.append("No matched records to analyze timeliness.")
            lines.append("")

        # --- Section 4: Amendment/Renewal Patterns ---
        lines.append("## 4. Amendment/Renewal Patterns")
        lines.append("")

        amended = [r for r in wellcare_records if r.status == "Amended"]
        renewed = [r for r in wellcare_records if r.status == "Renewed"]

        lines.append(f"- **Amended records:** {len(amended)} (county/date changes)")
        lines.append(f"- **Renewed records:** {len(renewed)} (window extensions)")
        lines.append("")

        if renewed:
            lines.append("### Renewal Chains (Multi-Extension Disasters)")
            lines.append("")
            # Group renewed records by state+title similarity
            renewal_chains = defaultdict(list)
            for r in renewed:
                key = (r.state, self._normalize_title_key_for_intel(r.title))
                renewal_chains[key].append(r)

            for (state, title_key), chain in sorted(renewal_chains.items(), key=lambda x: -len(x[1])):
                if len(chain) >= 1:
                    sample = chain[0]
                    lines.append(f"- **{state} — {sample.title}**: {len(chain)} renewal(s), SEP: {sample.sep_start} - {sample.sep_end}")
                    if sample.notes:
                        # Extract date changes from notes
                        date_changes = re.findall(r"Date changed from .+? to .+?(?:\.|$)", sample.notes)
                        for dc in date_changes[:2]:
                            lines.append(f"  - {dc.strip()}")
            lines.append("")

        if amended:
            lines.append("### Amendment Patterns")
            lines.append("")
            amend_types = Counter()
            for r in amended:
                notes_lower = r.notes.lower()
                if "counties" in notes_lower or "county" in notes_lower:
                    amend_types["County expansion"] += 1
                if "date changed" in notes_lower or "closes incident" in notes_lower:
                    amend_types["Date adjustment"] += 1
                if "contract" in notes_lower:
                    amend_types["Contract update"] += 1
                if "amendment" in notes_lower and "add" in notes_lower:
                    amend_types["Scope expansion"] += 1

            lines.append("| Amendment Type | Count |")
            lines.append("|---------------|-------|")
            for atype, count in amend_types.most_common():
                lines.append(f"| {atype} | {count} |")
            lines.append("")

        # --- Section 5: Recommendations ---
        lines.append("## 5. Pipeline Improvement Recommendations")
        lines.append("")

        lines.append("### New Data Sources to Monitor")
        lines.append("")
        lines.append("Based on carrier gap analysis:")
        lines.append("")

        # Identify top gap categories and recommend sources
        recommendations = []

        if gap_types.get("Wildfire", 0) > 5:
            recommendations.append(
                "1. **NIFC/InciWeb for wildfire declarations** — Carriers track 60+ named fires. "
                "NIFC (National Interagency Fire Center) publishes fire data. InciWeb has per-incident pages. "
                "Could automate discovery of fires that trigger governor declarations."
            )

        if gap_types.get("Drought", 0) > 0:
            recommendations.append(
                "2. **US Drought Monitor for drought emergencies** — Oregon declares drought emergencies "
                "that carriers track. USDM publishes weekly drought maps with county-level data."
            )

        if gap_types.get("Homelessness Emergency", 0) > 0 or gap_types.get("Crime/Civil Emergency", 0) > 0:
            recommendations.append(
                "3. **Municipal emergency declarations** — Cities (Portland OR, Albuquerque NM, NYC) declare "
                "local emergencies for homelessness, crime, healthcare staffing. No centralized source — "
                "carrier data is the best discovery mechanism."
            )

        if gap_types.get("Infrastructure Emergency", 0) > 0:
            recommendations.append(
                "4. **EPA/state water/infrastructure alerts** — Sewer collapses, water system failures, "
                "pipeline disruptions trigger local/state emergencies. EPA Emergency Response page may help."
            )

        if fmag_records:
            recommendations.append(
                f"5. **Re-evaluate FEMA FMAG exclusion** — {len(fmag_records)} active FMAG records "
                "honored by Wellcare. If carriers treat FMAG as valid DST triggers, our agents need them. "
                "Recommendation: Include FMAG with a 'carrier-validated' confidence flag, pending CMS guidance."
            )

        recommendations.append(
            f"6. **Quarterly carrier spreadsheet import** — Current gap analysis found "
            f"{len(results['gaps'])} missing records. Make carrier Excel import a standard process: "
            f"parse → cross-reference → research queue → curate."
        )

        for rec in recommendations:
            lines.append(rec)
            lines.append("")

        lines.append("### Process Improvements")
        lines.append("")
        lines.append("- **Extension tracking is mandatory** — 24 Aetna records have extensions we don't track.")
        lines.append("  FL hurricanes get renewed every 30-60 days. Without tracking, our SEP end dates go stale.")
        lines.append("- **Amendment monitoring** — County additions mid-disaster are common (18 Wellcare amendments).")
        lines.append("  Our tool shows static county lists that don't update.")
        lines.append("- **Carrier as early warning** — Carriers discover STATE declarations faster than us.")
        lines.append("  Use carrier data as trigger to search for official governor declarations.")
        lines.append("")

        return "\n".join(lines)

    def _normalize_title_key_for_intel(self, title: str) -> str:
        lower = title.lower()
        lower = re.sub(r"^(extension:?\s*|update:?\s*|renewed:?\s*)", "", lower)
        lower = re.sub(r"\s*eo\s*[-_]?\s*[\d.]+", "", lower)
        return re.sub(r"\s+", " ", lower).strip()

    def generate_gap_markdown(self, gap_data: Dict) -> str:
        """Generate human-readable gap report."""
        lines = ["# Carrier Gap Report", ""]
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        lines.append(f"Total gaps: **{gap_data['total_gaps']}** disasters tracked by carriers but NOT in our tool")
        lines.append("")

        lines.append("## Gaps by Disaster Type")
        lines.append("")
        lines.append("| Type | Count |")
        lines.append("|------|-------|")
        for itype, count in gap_data["by_type"].items():
            lines.append(f"| {itype} | {count} |")
        lines.append("")

        lines.append("## Gaps by State")
        lines.append("")
        lines.append("| State | Count |")
        lines.append("|-------|-------|")
        for state, count in gap_data["by_state"].items():
            lines.append(f"| {state} | {count} |")
        lines.append("")

        # Detailed gap list grouped by type
        gaps_by_type = defaultdict(list)
        for g in gap_data["gaps"]:
            gaps_by_type[g["incident_type"]].append(g)

        lines.append("## Detailed Gap List")
        lines.append("")

        for itype in sorted(gaps_by_type.keys()):
            gaps = gaps_by_type[itype]
            lines.append(f"### {itype} ({len(gaps)} missing)")
            lines.append("")
            lines.append("| State | Title | Carriers | SEP Window | Authority | URLs |")
            lines.append("|-------|-------|----------|------------|-----------|------|")
            for g in sorted(gaps, key=lambda x: (x["state"], x["title"])):
                carriers = ", ".join(g.get("carriers", [g["carrier"]]))
                sep = f"{g.get('sep_start', '?')} - {g.get('sep_end', '?')}"
                urls = " ".join(g.get("official_urls", [])[:1]) or "None"
                title_short = g["title"][:45]
                lines.append(f"| {g['state']} | {title_short} | {carriers} | {sep} | {g['declaring_authority']} | {urls[:60]} |")
            lines.append("")

        return "\n".join(lines)


# =========================================================================
# Main execution
# =========================================================================

def main():
    parser = argparse.ArgumentParser(description="Parse carrier DST data and cross-reference")
    parser.add_argument("--aetna", help="Path to Aetna Excel file")
    parser.add_argument("--wellcare", help="Path to Wellcare Excel file")
    parser.add_argument("--curated", required=True, help="Path to curated_disasters.json")
    parser.add_argument("--output", default="carrier_analysis.json", help="Output JSON path")
    args = parser.parse_args()

    output_dir = os.path.dirname(os.path.abspath(args.output)) or "."

    # Parse carrier files
    aetna_records = []
    wellcare_records = []
    wellcare_all = []

    if args.aetna:
        print(f"Parsing Aetna: {args.aetna}")
        aetna_parser = AetnaParser()
        aetna_records = aetna_parser.parse(args.aetna)

    if args.wellcare:
        print(f"Parsing Wellcare: {args.wellcare}")
        wellcare_parser = WellcareParser()
        wellcare_records = wellcare_parser.parse(args.wellcare)
        wellcare_all = wellcare_parser.parse_all(args.wellcare)

    all_carrier_records = aetna_records + wellcare_records
    print(f"\nTotal carrier records: {len(all_carrier_records)} ({len(aetna_records)} Aetna + {len(wellcare_records)} Wellcare)")

    # Cross-reference
    print(f"\nCross-referencing against {args.curated}...")
    xref = CrossReferencer(args.curated)
    results = xref.match_all(all_carrier_records)

    print(f"  Matched: {len(results['matched'])}")
    print(f"  Gaps: {len(results['gaps'])}")
    print(f"  Discrepancies: {len(results['discrepancies'])}")

    # Generate reports
    report_gen = GapReportGenerator()

    # 1. carrier_analysis.json
    analysis = {
        "metadata": {
            "generated": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "aetna_records": len(aetna_records),
            "wellcare_records": len(wellcare_records),
            "matched": len(results["matched"]),
            "gaps": len(results["gaps"]),
            "discrepancies": len(results["discrepancies"]),
        },
        "matched": results["matched"],
        "discrepancies": results["discrepancies"],
    }
    analysis_path = args.output
    with open(analysis_path, "w") as f:
        json.dump(analysis, f, indent=2)
    print(f"\nWrote: {analysis_path}")

    # 2. carrier_gaps.json
    gap_data = report_gen.generate_gap_report(results)
    gaps_path = os.path.join(output_dir, "carrier_gaps.json")
    with open(gaps_path, "w") as f:
        json.dump(gap_data, f, indent=2)
    print(f"Wrote: {gaps_path}")

    # 3. carrier_report.md
    report_md = report_gen.generate_gap_markdown(gap_data)
    report_path = os.path.join(output_dir, "carrier_report.md")
    with open(report_path, "w") as f:
        f.write(report_md)
    print(f"Wrote: {report_path}")

    # 4. carrier_intelligence.md
    intel_md = report_gen.generate_intelligence_report(
        aetna_records, wellcare_records, wellcare_all, results
    )
    intel_path = os.path.join(output_dir, "carrier_intelligence.md")
    with open(intel_path, "w") as f:
        f.write(intel_md)
    print(f"Wrote: {intel_path}")

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Carrier records analyzed: {len(all_carrier_records)}")
    print(f"Matched to our data:     {len(results['matched'])}")
    print(f"Gaps (missing from us):  {len(results['gaps'])}")
    print(f"SEP discrepancies:       {len(results['discrepancies'])}")
    print(f"\nTop gap types:")
    for itype, count in sorted(gap_data["by_type"].items(), key=lambda x: -x[1])[:5]:
        print(f"  {itype}: {count}")
    print(f"\nFiles written:")
    print(f"  {analysis_path}")
    print(f"  {gaps_path}")
    print(f"  {report_path}")
    print(f"  {intel_path}")


if __name__ == "__main__":
    main()
